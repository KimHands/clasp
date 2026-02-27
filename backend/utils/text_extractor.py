import os
from typing import Optional


def extract_text(file_path: str) -> Optional[str]:
    """
    파일 확장자에 따라 텍스트 추출 전략 선택
    - PDF: 1~2페이지 스킵 후 4구간 샘플링 (30/45/65/85%)
    - DOCX: 단락 기반 추출
    - DOC: textutil(macOS) / antiword 사용
    - TXT/MD: 전체 읽기 (최대 5000자)
    - XLSX: 첫 시트 헤더 + 앞 5행 (openpyxl)
    - CSV: 헤더 + 앞 5행 (내장 csv 모듈)
    """
    ext = os.path.splitext(file_path)[1].lower()
    try:
        if ext == ".pdf":
            return _extract_pdf(file_path)
        elif ext == ".docx":
            return _extract_docx(file_path)
        elif ext == ".doc":
            return _extract_doc(file_path)
        elif ext in (".txt", ".md"):
            return _extract_plain(file_path)
        elif ext == ".xlsx":
            return _extract_xlsx(file_path)
        elif ext == ".csv":
            return _extract_csv(file_path)
    except Exception:
        return None
    return None


def _extract_pdf(file_path: str) -> Optional[str]:
    try:
        import fitz  # PyMuPDF
    except ImportError:
        return None

    doc = None
    try:
        doc = fitz.open(file_path)
        total_pages = len(doc)

        if total_pages == 0:
            return None

        # 3페이지 이상인 경우 1~2페이지(표지/목차) 스킵
        start_page = 2 if total_pages >= 3 else 0
        effective_pages = list(range(start_page, total_pages))

        if not effective_pages:
            return None

        # 유효 페이지가 4 미만이면 샘플링 없이 전체 추출 (중복 샘플링 방지)
        if len(effective_pages) < 4:
            chunks = []
            for page_idx in effective_pages:
                text = doc[page_idx].get_text("text")
                if text:
                    chunks.append(text[:1200])
            return "\n".join(chunks) if chunks else None

        # 4구간 샘플링 위치 (30%, 45%, 65%, 85%)
        sample_ratios = [0.30, 0.45, 0.65, 0.85]
        n = len(effective_pages)
        sampled_indices = {effective_pages[min(int(r * n), n - 1)] for r in sample_ratios}

        chunks = []
        for page_idx in sorted(sampled_indices):
            page = doc[page_idx]
            text = page.get_text("text")
            if text:
                chunks.append(text[:300])

        return "\n".join(chunks) if chunks else None
    finally:
        if doc:
            doc.close()


def _extract_docx(file_path: str) -> Optional[str]:
    try:
        from docx import Document
    except ImportError:
        return None

    doc = Document(file_path)
    paragraphs = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
    return "\n".join(paragraphs)[:5000] if paragraphs else None


def _extract_doc(file_path: str) -> Optional[str]:
    """
    .doc (레거시 Word) 파일 텍스트 추출.
    subprocess로 textutil(macOS) 또는 antiword를 시도.
    """
    import subprocess
    import platform

    if platform.system() == "Darwin":
        try:
            result = subprocess.run(
                ["textutil", "-convert", "txt", "-stdout", file_path],
                capture_output=True, text=True, timeout=10,
            )
            if result.returncode == 0 and result.stdout.strip():
                return result.stdout.strip()[:5000]
        except (FileNotFoundError, subprocess.TimeoutExpired):
            pass

    try:
        result = subprocess.run(
            ["antiword", file_path],
            capture_output=True, text=True, timeout=10,
        )
        if result.returncode == 0 and result.stdout.strip():
            return result.stdout.strip()[:5000]
    except (FileNotFoundError, subprocess.TimeoutExpired):
        pass

    return None


def _extract_plain(file_path: str) -> Optional[str]:
    try:
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            return f.read(5000)
    except Exception:
        return None


def _extract_xlsx(file_path: str) -> Optional[str]:
    """
    XLSX 첫 번째 시트에서 헤더 행 + 앞 5행 텍스트 추출.
    열 이름과 셀 값을 쉼표로 연결해 Tier 2 임베딩 입력으로 활용.
    read_only=True + data_only=True 로 수식 대신 계산값, 빠른 읽기 보장.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None

    wb = None
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 6:  # 헤더 1행 + 데이터 5행
                break
            cells = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
            if cells:
                rows.append(", ".join(cells))
        return "\n".join(rows)[:5000] if rows else None
    finally:
        if wb:
            wb.close()


def _extract_csv(file_path: str) -> Optional[str]:
    """
    CSV 헤더 행 + 앞 5행 텍스트 추출.
    열 이름이 내용 분류에 가장 유용한 정보를 담고 있으므로 헤더를 우선 포함.
    """
    import csv

    rows = []
    for encoding in ("utf-8", "utf-8-sig", "cp949", "euc-kr"):
        try:
            with open(file_path, "r", encoding=encoding, errors="strict", newline="") as f:
                reader = csv.reader(f)
                for i, row in enumerate(reader):
                    if i >= 6:  # 헤더 1행 + 데이터 5행
                        break
                    cells = [cell.strip() for cell in row if cell.strip()]
                    if cells:
                        rows.append(", ".join(cells))
            break  # 인코딩 성공 시 루프 탈출
        except (UnicodeDecodeError, UnicodeError):
            rows = []
            continue
        except Exception:
            return None

    return "\n".join(rows)[:5000] if rows else None
